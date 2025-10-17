from flask import Flask, render_template, request, redirect
import sqlite3
import os
from openpyxl import load_workbook

app = Flask(__name__)
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(BASE_DIR, "inventory.db")
UPLOAD_FOLDER = os.path.join(BASE_DIR, "static", "reports")

if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

def init_db():
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute('''CREATE TABLE IF NOT EXISTS inventory (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        tool_type TEXT,
        serial_number TEXT,
        size TEXT,
        thread_type TEXT,
        location TEXT,
        status TEXT,
        report_link TEXT,
        description TEXT
    )''')
    conn.commit()
    conn.close()

init_db()

@app.route("/", methods=["GET"])
def index():
    query = "SELECT * FROM inventory WHERE 1=1"
    params = []
    tool_type = request.args.get("tool_type", "")
    serial_number = request.args.get("serial_number", "")
    status = request.args.get("status", "")
    location = request.args.get("location", "")

    if tool_type:
        query += " AND tool_type LIKE ?"
        params.append(f"%{tool_type}%")
    if serial_number:
        query += " AND serial_number LIKE ?"
        params.append(f"%{serial_number}%")
    if status:
        query += " AND status LIKE ?"
        params.append(f"%{status}%")
    if location:
        query += " AND location LIKE ?"
        params.append(f"%{location}%")

    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    c.execute(query, params)
    tools = c.fetchall()
    conn.close()
    return render_template("index.html", tools=tools)

@app.route("/add", methods=["POST"])
def add():
    tool_type = request.form.get("tool_type")
    serial_number = request.form.get("serial_number")
    size = request.form.get("size")
    thread_type = request.form.get("thread_type")
    location = request.form.get("location")
    status = request.form.get("status")
    description = request.form.get("description")
    report_file = request.files.get("report_file")

    report_link = ""
    if report_file and report_file.filename:
        save_path = os.path.join(UPLOAD_FOLDER, report_file.filename)
        report_file.save(save_path)
        report_link = "/static/reports/" + report_file.filename

    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("SELECT id FROM inventory WHERE serial_number=?", (serial_number,))
    if c.fetchone():
        conn.close()
        return "<script>alert('شماره سریال تکراری است'); window.location.href='/';</script>"

    c.execute('''INSERT INTO inventory
        (tool_type, serial_number, size, thread_type, location, status, report_link, description)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)''',
        (tool_type, serial_number, size, thread_type, location, status, report_link, description))
    conn.commit()
    conn.close()
    return redirect("/")

@app.route("/edit/<int:id>", methods=["GET", "POST"])
def edit(id):
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    c = conn.cursor()

    if request.method == "POST":
        tool_type = request.form.get("tool_type")
        serial_number = request.form.get("serial_number")
        size = request.form.get("size")
        thread_type = request.form.get("thread_type")
        location = request.form.get("location")
        status = request.form.get("status")
        description = request.form.get("description")

        c.execute('''UPDATE inventory SET
            tool_type=?, serial_number=?, size=?, thread_type=?, location=?, status=?, description=?
            WHERE id=?''',
            (tool_type, serial_number, size, thread_type, location, status, description, id))
        conn.commit()
        conn.close()
        return redirect("/")

    c.execute("SELECT * FROM inventory WHERE id=?", (id,))
    item = c.fetchone()
    conn.close()
    return render_template("edit.html", item=item)

@app.route("/delete/<int:id>")
def delete(id):
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("DELETE FROM inventory WHERE id=?", (id,))
    conn.commit()
    conn.close()
    return redirect("/")

@app.route("/upload_excel", methods=["POST"])
def upload_excel():
    excel_file = request.files.get("file")
    if not excel_file or not excel_file.filename.endswith(".xlsx"):
        return "<script>alert('لطفاً فایل Excel معتبر انتخاب کنید.'); window.location.href='/';</script>"

    wb = load_workbook(excel_file)
    sheet = wb.active
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    skipped = []

    for idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        if idx == 1: continue
        if not row[0]: continue
        tool_type, serial_number, size, thread_type, location, status = row[:6]

        # بررسی رکورد تکراری بر اساس شماره سریال
        c.execute("SELECT id FROM inventory WHERE serial_number=?", (serial_number,))
        if c.fetchone():
            skipped.append(serial_number)
            continue

        c.execute('''INSERT INTO inventory
            (tool_type, serial_number, size, thread_type, location, status, report_link, description)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)''',
            (tool_type, serial_number, size, thread_type, location, status, "", ""))

    conn.commit()
    conn.close()

    if skipped:
        return "<script>alert('شماره سریال‌های تکراری نادیده گرفته شدند: {}'); window.location.href='/';</script>".format(', '.join(skipped))

    return redirect("/")

@app.route("/update_description/<int:id>", methods=["POST"])
def update_description(id):
    description = request.form.get("description", "")
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("UPDATE inventory SET description=? WHERE id=?", (description, id))
    conn.commit()
    conn.close()
    return "OK"

@app.route("/delete_selected", methods=["POST"])
def delete_selected():
    ids = request.form.getlist('ids')
    if ids:
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        c.executemany("DELETE FROM inventory WHERE id=?", [(i,) for i in ids])
        conn.commit()
        conn.close()
    return '', 204

@app.route("/delete_all_filtered", methods=["POST"])
def delete_all_filtered():
    tool_type = request.form.get("tool_type", "")
    serial_number = request.form.get("serial_number", "")
    status = request.form.get("status", "")
    location = request.form.get("location", "")
    query = "DELETE FROM inventory WHERE 1=1"
    params = []

    if tool_type:
        query += " AND tool_type LIKE ?"
        params.append(f"%{tool_type}%")
    if serial_number:
        query += " AND serial_number LIKE ?"
        params.append(f"%{serial_number}%")
    if status:
        query += " AND status LIKE ?"
        params.append(f"%{status}%")
    if location:
        query += " AND location LIKE ?"
        params.append(f"%{location}%")

    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute(query, params)
    conn.commit()
    conn.close()
    return '', 204

if __name__ == "__main__":
    app.run(debug=True, host="0.0.0.0", port=5000)
